import openpyxl
import tkinter as tk
from tkinter import messagebox, filedialog, colorchooser, Toplevel
import os
import sys
import platform

def find_and_update(file_path, search_value, add_hours, record_id, history_text, entry_id, entry_hours, entry_record):
    wb = openpyxl.load_workbook(file_path)
    
    for sheet in wb.worksheets:
        for row in sheet.iter_rows(min_row=2, max_row=sheet.max_row, values_only=False):
            student_id = str(row[2].value).strip()  # C欄學號（轉為字串）
            student_name = str(row[3].value).strip()  # D欄姓名
            
            if search_value.strip() in [student_id, student_name]:
                # 更新 F 欄（時數累加）
                current_hours = row[5].value or 0
                row[5].value = current_hours + add_hours
                
                # 尋找 J 欄以後的空白欄位，寫入登記編號
                col_index = 9  # J 欄的索引是 9（從 0 計算）
                while col_index < sheet.max_column and row[col_index].value:
                    col_index += 1
                
                # 如果欄位不夠，自動增加
                if col_index >= sheet.max_column:
                    sheet.insert_cols(col_index + 1)
                row[col_index].value = record_id
                
                wb.save(file_path)
                wb.close()
                messagebox.showinfo("成功", f"更新成功：{search_value}\n時數 +{add_hours}\n登記編號 {record_id}")
                
                # 記錄歷史紀錄
                history_text.insert(tk.END, f"{search_value} | +{add_hours} 小時 | 編號: {record_id}\n")
                history_text.yview(tk.END)
                
                # 清空輸入欄位
                entry_id.delete(0, tk.END)
                entry_hours.delete(0, tk.END)
                entry_record.delete(0, tk.END)
                return
    
    wb.close()
    messagebox.showwarning("錯誤", "未找到符合條件的學生。")

def select_file(label_file):
    file_path = filedialog.askopenfilename(filetypes=[("Excel files", "*.xlsx;*.xlsm")])
    if file_path:
        label_file.config(text=os.path.basename(file_path))
        return file_path
    return None

def open_settings(root, font_size_var, widgets):
    settings_win = Toplevel(root)
    settings_win.title("設定")
    settings_win.geometry("400x300")
    
    tk.Label(settings_win, text="文字大小:", font=("標楷體", 12)).pack()
    scale = tk.Scale(settings_win, from_=8, to=20, orient=tk.HORIZONTAL, variable=font_size_var)
    scale.pack()
    
    def apply_font_size():
        new_size = font_size_var.get()
        for widget in widgets:
            widget.config(font=("標楷體", new_size))
    
    tk.Button(settings_win, text="套用", command=apply_font_size, font=("標楷體", 12)).pack()
    
    tk.Label(settings_win, text="背景顏色:", font=("標楷體", 12)).pack()
    tk.Button(settings_win, text="選擇顏色", command=lambda: change_color(root), font=("標楷體", 12)).pack()

def change_color(root):
    color = colorchooser.askcolor()[1]
    if color:
        root.configure(bg=color)

def focus_next_widget(event):
    event.widget.tk_focusNext().focus()
    return "break"

def focus_prev_widget(event):
    event.widget.tk_focusPrev().focus()
    return "break"

def main():
    root = tk.Tk()
    root.title("學生時數登記系統")
    root.geometry("800x600")  # 提高介面解析度
    root.resizable(True, True)  # 允許視窗自由縮放
    
    font_size = tk.IntVar(value=12)
    
    widgets = []
    
    # 設置功能按鈕
    top_frame = tk.Frame(root)
    top_frame.grid(row=0, column=0, columnspan=3, sticky="w")
    
    label_file = tk.Label(top_frame, text="未選擇檔案", font=("標楷體", 12))
    label_file.pack(side="left", padx=10)
    
    tk.Button(top_frame, text="開啟檔案", command=lambda: select_file(label_file), font=("標楷體", 12)).pack(side="left", padx=5)
    tk.Button(top_frame, text="設定", command=lambda: open_settings(root, font_size, widgets), font=("標楷體", 12)).pack(side="left", padx=5)
    
    labels = ["學號或姓名:", "增加的時數:", "登記編號:", "歷史紀錄:"]
    entries = []
    
    for i, text in enumerate(labels[:-1]):
        label = tk.Label(root, text=text, font=("標楷體", 12))
        label.grid(row=i+1, column=0)
        widgets.append(label)
        entry = tk.Entry(root, font=("標楷體", 12))
        entry.grid(row=i+1, column=1)
        entry.bind("<Return>", focus_next_widget)
        entry.bind("<Down>", focus_next_widget)
        entry.bind("<Up>", focus_prev_widget)
        entries.append(entry)
        widgets.append(entry)
    
    entry_id, entry_hours, entry_record = entries
    
    submit_button = tk.Button(root, text="提交", font=("標楷體", 12), 
                              command=lambda: find_and_update(label_file.cget("text"), entry_id.get(), int(entry_hours.get()), entry_record.get(), history_text, entry_id, entry_hours, entry_record))
    submit_button.grid(row=4, column=1, sticky="w")  # 提交按鈕移到登記編號輸入框的下方
    widgets.append(submit_button)
    
    history_label = tk.Label(root, text=labels[-1], font=("標楷體", 12))
    history_label.grid(row=5, column=0, columnspan=3)
    widgets.append(history_label)
    
    history_text = tk.Text(root, height=15, width=60, font=("標楷體", 12))
    history_text.grid(row=6, column=0, columnspan=3, sticky="nsew")
    widgets.append(history_text)
    
    root.columnconfigure(1, weight=1)
    root.rowconfigure(6, weight=1)
    
    root.mainloop()
    
if __name__ == "__main__":
    main()
